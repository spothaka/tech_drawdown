# --- project-relative paths (generalized from the scratch build workspace) ---
# Set TDD_BASE to the project root, or leave blank to auto-detect (parent of scripts/).
import os
TDD_BASE = os.environ.get("TDD_BASE") or os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR    = os.path.join(TDD_BASE, "data")
DASH_DIR    = os.path.join(TDD_BASE, "dashboard")
DOCS_DIR    = os.path.join(TDD_BASE, "docs")
PORT_DIR    = os.path.join(TDD_BASE, "portfolio")
SCRIPTS_DIR = os.path.join(TDD_BASE, "scripts")
# ---------------------------------------------------------------------------

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment

SRC=os.path.join(DATA_DIR,'Brokerage.xlsx')
OUT=os.path.join(PORT_DIR,'Brokerage_Portfolio.xlsx')

src=openpyxl.load_workbook(SRC,data_only=True).active
data=[]
for row in src.iter_rows(min_row=2,values_only=True):
    if row and row[0] and str(row[0]).strip():
        sym,qty,paid,val=(list(row)+[None,None,None,None])[:4]
        data.append((str(sym).strip(),qty,paid,val))
print("holdings read:",len(data))

GOLD="E0A106"; DKGOLD="B07D04"; INK="1F2937"; MUT="6B7280"
YEL="FFFBEA"; GREEN="D4EDDA"; AMBER="FFF3CD"; RED="F8D7DA"; LINEC="D9D9D9"
thin=Side(style="thin",color=LINEC); border=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook(); ws=wb.active; ws.title="Brokerage Portfolio"; ws.sheet_view.showGridLines=False

HDR=["Symbol","Quantity","Price Paid $","Value $ (import)","Asset Type","Cost Basis $",
     "Current Price ($)","52-Wk High ($)","% Off High","Status",
     "200-Day SMA ($)","% vs 200d","50-Day SMA ($)","% vs 50d","Cross Signal",
     "Market Value ($)","Gain/Loss ($)","Gain/Loss %","Weight %"]
NCOL=len(HDR)  # 19 -> A..S
HROW=8; FIRST=9; LAST=48
COLW=[30,10,12,15,12,13,12,12,9,11,12,9,12,9,13,14,13,10,8]
for i,w in enumerate(COLW,1): ws.column_dimensions[chr(64+i)].width=w

def L(i): return chr(64+i)
lastL=L(NCOL)

# title + note
ws.merge_cells(f"A1:{lastL}1")
c=ws["A1"]; c.value="Brokerage Portfolio  —  Private"; c.font=Font(name="Arial",size=18,bold=True,color="FFFFFF")
c.fill=PatternFill("solid",fgColor=GOLD); c.alignment=Alignment(horizontal="left",vertical="center",indent=1); ws.row_dimensions[1].height=30
ws.merge_cells(f"A2:{lastL}2")
c=ws["A2"]; c.value=("Live metrics auto-pull via STOCKHISTORY (Excel 365 required). Edit only the shaded broker columns A–D "
 "(Symbol, Quantity, Price Paid, Value) — paste a fresh export over them, then press F9 / Data ▸ Refresh All.")
c.font=Font(name="Arial",size=9,italic=True,color=MUT); c.alignment=Alignment(horizontal="left",indent=1)

def lbl(cell,t): ws[cell]=t; ws[cell].font=Font(name="Arial",size=10,bold=True,color=INK)
def val(cell,f,fmt): ws[cell]=f; ws[cell].font=Font(name="Arial",size=10,bold=True,color=DKGOLD); ws[cell].number_format=fmt
lbl("A3","Total Market Value"); val("C3",f"=SUM(P{FIRST}:P{LAST})",'$#,##0.00')
lbl("A4","Total Cost Basis");   val("C4",f"=SUM(F{FIRST}:F{LAST})",'$#,##0.00')
lbl("A5","Total Gain / Loss");  val("C5",f"=IF(SUM(F{FIRST}:F{LAST})=0,\"\",SUM(Q{FIRST}:Q{LAST}))",'$#,##0.00;($#,##0.00)')
lbl("E5","Gain / Loss %");      val("G5",f"=IFERROR(SUM(Q{FIRST}:Q{LAST})/SUM(F{FIRST}:F{LAST}),\"\")",'0.0%')
lbl("A6","Holdings");           val("C6",f"=COUNTA(A{FIRST}:A{LAST})",'0')

for j,h in enumerate(HDR,1):
    cell=ws.cell(row=HROW,column=j,value=h)
    cell.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    cell.fill=PatternFill("solid",fgColor=DKGOLD)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=border
ws.row_dimensions[HROW].height=30

for r in range(FIRST,LAST+1):
    M=f'OR($E{r}="Equity/ETF",$E{r}="Mutual Fund")'
    ws[f"E{r}"]=(f'=IF($A{r}="","",IF(ISNUMBER(SEARCH("CD",$A{r})),"CD",IF(UPPER($A{r})="CASH","Cash",'
                 f'IF(AND(LEN($A{r})=5,RIGHT(UPPER($A{r}),1)="X"),"Mutual Fund","Equity/ETF"))))')
    ws[f"F{r}"]=(f'=IF($A{r}="","",IF($E{r}="CD",IFERROR($B{r}*$C{r}/100,""),'
                 f'IF(AND(ISNUMBER($B{r}),ISNUMBER($C{r})),$B{r}*$C{r},"")))')
    ws[f"G{r}"]=(f'=IF(OR($A{r}="",NOT({M})),"",IFERROR(INDEX(_xlfn.STOCKHISTORY($A{r},TODAY()-7,TODAY(),0,0,0,1),'
                 f'ROWS(_xlfn.STOCKHISTORY($A{r},TODAY()-7,TODAY(),0,0,0,1)),2),""))')
    ws[f"H{r}"]=(f'=IF(OR($A{r}="",NOT({M})),"",IFERROR(MAX(_xlfn.STOCKHISTORY($A{r},TODAY()-365,TODAY(),0,0,3)),""))')
    ws[f"I{r}"]=f'=IF(OR($G{r}="",$H{r}=""),"",IFERROR(($G{r}-$H{r})/$H{r},""))'
    ws[f"J{r}"]=f'=IF($I{r}="","",IF($I{r}<=-0.2,"Bear",IF($I{r}<=-0.1,"Correction","Normal")))'
    ws[f"K{r}"]=(f'=IF(OR($A{r}="",NOT({M})),"",IFERROR(AVERAGE(_xlfn.STOCKHISTORY($A{r},TODAY()-290,TODAY(),0,0,1)),""))')
    ws[f"L{r}"]=f'=IF(OR($G{r}="",$K{r}=""),"",IFERROR($G{r}/$K{r}-1,""))'
    ws[f"M{r}"]=(f'=IF(OR($A{r}="",NOT({M})),"",IFERROR(AVERAGE(_xlfn.STOCKHISTORY($A{r},TODAY()-73,TODAY(),0,0,1)),""))')
    ws[f"N{r}"]=f'=IF(OR($G{r}="",$M{r}=""),"",IFERROR($G{r}/$M{r}-1,""))'
    ws[f"O{r}"]=f'=IF(OR($K{r}="",$M{r}=""),"",IF($M{r}>=$K{r},"Golden Cross","Death Cross"))'
    ws[f"P{r}"]=(f'=IF($A{r}="","",IF({M},IF(ISNUMBER($G{r}),$B{r}*$G{r},$D{r}),$D{r}))')
    ws[f"Q{r}"]=f'=IF(OR($P{r}="",NOT(ISNUMBER($F{r}))),"",$P{r}-$F{r})'
    ws[f"R{r}"]=f'=IF(OR($Q{r}="",NOT(ISNUMBER($F{r})),$F{r}=0),"",$Q{r}/$F{r})'
    ws[f"S{r}"]=f'=IF(OR($P{r}="",SUM($P${FIRST}:$P${LAST})=0),"",$P{r}/SUM($P${FIRST}:$P${LAST}))'
    for col in ("C","D","F","G","H","K","M","P"): ws[f"{col}{r}"].number_format='$#,##0.00'
    ws[f"Q{r}"].number_format='$#,##0.00;($#,##0.00)'
    for col in ("I","L","N","R","S"): ws[f"{col}{r}"].number_format='0.0%'
    ws[f"B{r}"].number_format='#,##0.####'
    for col in ("A","B","C","D"):
        ws[f"{col}{r}"].fill=PatternFill("solid",fgColor=YEL); ws[f"{col}{r}"].font=Font(name="Arial",size=10,color="0000FF")
    for j in range(1,NCOL+1):
        cell=ws.cell(row=r,column=j); cell.border=border
        if L(j) not in ("A","B","C","D"): cell.font=Font(name="Arial",size=10,color=INK)
        if L(j) in ("E","J","O"): cell.alignment=Alignment(horizontal="center")

# fill broker data A-D
for i,(sym,qty,paid,vv) in enumerate(data):
    r=FIRST+i
    ws[f"A{r}"]=sym; ws[f"B{r}"]=qty; ws[f"C{r}"]=paid; ws[f"D{r}"]=vv

def cf(col,word,fill):
    ws.conditional_formatting.add(f"{col}{FIRST}:{col}{LAST}",
        CellIsRule(operator="equal",formula=[f'"{word}"'],fill=PatternFill("solid",fgColor=fill)))
cf("J","Bear",RED); cf("J","Correction",AMBER); cf("J","Normal",GREEN)
cf("O","Golden Cross",GREEN); cf("O","Death Cross",RED)
ws.conditional_formatting.add(f"Q{FIRST}:R{LAST}",CellIsRule(operator="lessThan",formula=["0"],font=Font(color="C00000")))
ws.conditional_formatting.add(f"Q{FIRST}:R{LAST}",CellIsRule(operator="greaterThan",formula=["0"],font=Font(color="1E7D32")))

ws[f"C{HROW}"].comment=Comment("Avg cost per share (equities/funds). For CDs this is price per 100 par.","IRA")
ws[f"D{HROW}"].comment=Comment("Broker snapshot value. Used as market value for CDs/cash and as fallback if a live price is unavailable.","IRA")
ws.freeze_panes="A9"
wb.save(OUT)
print("saved",OUT)
