"""Refresh DCF base prices from data/market_data.xlsx -> data/dcf.json + data/dcf_wacc.json.

The universe tables reprice daily (FMP batch-quote -> market_data.xlsx) but the DCF bases only
reprice when build_dcf.py re-fetches a symbol — which the finite/plan-gated analyst+statements
endpoints may prevent for long stretches. Left alone, the popup's "trading X% above/below",
reverse DCF, and Monte Carlo P(undervalued) drift against a price the universe tab one click
away no longer shows. This transform closes that gap using data the pipeline already fetched —
no new connector calls.

Capital weights are recomputed from the refreshed price (wE = mcap/(mcap+debt), the same formula
as build_dcf._wacc_inputs) so WACC never mixes a market cap from one date with a price from
another. kdAT is left as stored: the cost of debt is an assumption, not a price-derived value,
and silently re-deriving it would move fair values for no data reason.

Wired into rebuild_all.py (before the dcf.json load); safe to run standalone:
    python scripts/refresh_dcf_prices.py
"""
import os, json

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
MKT = os.path.join(ROOT, 'data', 'market_data.xlsx')
DCF = os.path.join(ROOT, 'data', 'dcf.json')
WACC = os.path.join(ROOT, 'data', 'dcf_wacc.json')


def load_prices(tickers):
    """ticker -> price from market_data.xlsx (first sheet hit wins; sheets share tickers)."""
    import openpyxl
    wb = openpyxl.load_workbook(MKT, read_only=True, data_only=True)
    want = set(tickers)
    out = {}
    for sh in wb.sheetnames:
        for r in wb[sh].iter_rows(min_row=2, values_only=True):
            if r and r[0] in want and r[0] not in out and isinstance(r[1], (int, float)) and r[1] > 0:
                out[r[0]] = float(r[1])
        if len(out) == len(want):
            break
    wb.close()
    return out


def _weights(price, shares, total_debt):
    mcap = price * shares
    tot = mcap + (total_debt or 0)
    wE = (mcap / tot) if tot else 1.0
    return round(wE, 5), round(1.0 - wE, 5)


def main():
    dcf = json.load(open(DCF, encoding='utf-8')) if os.path.exists(DCF) else {}
    wacc = json.load(open(WACC, encoding='utf-8')) if os.path.exists(WACC) else {}
    px = load_prices(set(dcf) | set(wacc))
    lines = []

    for sym, b in dcf.items():
        p = px.get(sym)
        if not p:
            lines.append(f"{sym}: no market_data price — kept {b.get('price')}")
            continue
        old = b.get('price')
        b['price'] = p
        b['wE'], b['wD'] = _weights(p, b['shares'], b.get('totalDebt'))
        if old != p:
            lines.append(f"{sym}: price {old} -> {p}  (wE {b['wE']})")

    for sym, w in wacc.items():
        p = px.get(sym)
        if not p:
            continue
        w['price'] = p
        base = dcf.get(sym)
        if base and 'wE' in w:                      # keep the legacy wE/wD entries in sync
            w['wE'], w['wD'] = base['wE'], base['wD']

    json.dump(dcf, open(DCF, 'w', encoding='utf-8'), separators=(',', ':'))
    json.dump(wacc, open(WACC, 'w', encoding='utf-8'), indent=1)
    print('DCF price refresh: %d covered bases, %d staged entries' % (len(dcf), len(wacc)))
    for ln in lines:
        print(' ', ln)


if __name__ == '__main__':
    main()
