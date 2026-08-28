"""FMP-primary universe build (Phase 2 + Phase 4a assembler).
Pure transform: FMP batch-quote JSON -> canonical rows -> integrity_guard (band)
-> data/market_data.xlsx (live-data workbook) + parity vs STOCKHISTORY, and
--emit-data assembles full DATA universe rows (FMP live + workbook static).
Chain: FMP -> Bigdata (--extra) -> carry-forward -> STOCKHISTORY/guard."""
from __future__ import annotations
import json, sys, os, datetime
HERE = os.path.dirname(os.path.abspath(__file__))
if HERE not in sys.path: sys.path.insert(0, HERE)
import integrity_guard as ig
try:
    from log_util import ev
except Exception:
    def ev(*a, **k): pass
DATA_DIR = os.path.join(HERE, '..', 'data')
WB_DEFAULT = os.path.join(DATA_DIR, 'tech100_drawdown_SP_NAS.xlsx')
MD_DEFAULT = os.path.join(DATA_DIR, 'market_data.xlsx')
SHEET = {'sp':'S&P 500 Live','nasdaq':'Nasdaq-100 Live','dow':'Dow Jones 100 Live','etfs':'Top 100 ETFs Live','thematic':'Thematic ETFs Live','mutualfunds':'Mutual Funds Live'}
TABS = list(SHEET.keys()); FUND_KEYS = {'etfs','thematic','mutualfunds'}
HMAP = {'Ticker':'ticker','Company':'company','Sector':'sector','Category':'category','Current Price ($)':'price','52-Wk High ($)':'high','200-Day SMA ($)':'sma','50-Day SMA ($)':'sma50','Analyst Consensus':'consensus','Forward P/E':'fwdpe','Dividend Yield % (est.)':'divyield'}
OUT_COLS = ['ticker','price','high','low','sma','sma50','off','recover','smapct','sma50pct','status','signal','cross','as_of','source','dq']
TODAY = datetime.date.today().isoformat()
def _num(v): return ig._num(v)
def derive_row(tk, q):
    price=_num(q.get('price')); high=_num(q.get('yearHigh')); low=_num(q.get('yearLow')); sma=_num(q.get('priceAvg200')); sma50=_num(q.get('priceAvg50'))
    if price is None: return None
    r={'ticker':tk,'price':price,'high':high,'low':low,'sma':sma,'sma50':sma50,'as_of':TODAY,'source':'FMP','dq':None}
    if high:
        r['off']=round(price/high-1,6); r['recover']=round(high/price-1,6); o=r['off']
        r['status']='Bear' if o<=-0.2 else 'Correction' if o<=-0.1 else 'Normal'
        r['signal']=('Deep value zone (>50% off)' if o<=-0.5 else 'Bear (-20% to -50%)' if o<=-0.2 else 'Correction (-10% to -20%)' if o<=-0.1 else 'Near highs (<10% off)')
    if sma: r['smapct']=round(price/sma-1,6)
    if sma50: r['sma50pct']=round(price/sma50-1,6)
    if sma and sma50: r['cross']='Golden Cross' if sma50>=sma else 'Death Cross'
    return r
def read_wb_tab(wb, tab_key):
    ws=wb[SHEET[tab_key]]; rows=list(ws.iter_rows(values_only=True))
    hi=next(i for i,r in enumerate(rows) if r and any(str(c).strip()=='Ticker' for c in r if c))
    hdr=[HMAP.get(str(c).strip()) if c else None for c in rows[hi]]; out=[]
    for r in rows[hi+1:]:
        t=r[0]
        if not t or not str(t).strip(): continue
        if str(t).strip().lower().startswith('summary'): break
        rec={hdr[i]:r[i] for i in range(len(hdr)) if hdr[i]}; rec['ticker']=str(rec['ticker']).strip(); out.append(rec)
    return out
def load_prior(md_path):
    if not os.path.exists(md_path): return {}
    import openpyxl
    wb=openpyxl.load_workbook(md_path,data_only=True,read_only=True); prior={}
    for tab in TABS:
        if tab not in wb.sheetnames: continue
        ws=wb[tab]; rows=list(ws.iter_rows(values_only=True))
        if not rows: continue
        hdr=[str(c).strip() if c else None for c in rows[0]]; d={}
        for r in rows[1:]:
            rec={hdr[i]:r[i] for i in range(len(hdr)) if hdr[i]}
            if rec.get('ticker'): d[str(rec['ticker']).strip()]=rec
        prior[tab]=d
    return prior
def _derive(r):
    price=r.get('price'); high=r.get('high'); sma=r.get('sma'); sma50=r.get('sma50')
    if price and high:
        r['off']=round(price/high-1,6); r['recover']=round(high/price-1,6); o=r['off']
        r['status']='Bear' if o<=-0.2 else 'Correction' if o<=-0.1 else 'Normal'
        r['signal']=('Deep value zone (>50% off)' if o<=-0.5 else 'Bear (-20% to -50%)' if o<=-0.2 else 'Correction (-10% to -20%)' if o<=-0.1 else 'Near highs (<10% off)')
    if price and sma: r['smapct']=round(price/sma-1,6)
    if price and sma50: r['sma50pct']=round(price/sma50-1,6)
    if sma and sma50: r['cross']='Golden Cross' if sma50>=sma else 'Death Cross'
    return r
def derive_from_wb(tk, wr):
    r={'ticker':tk,'price':_num(wr.get('price')),'high':_num(wr.get('high')),'low':None,'sma':_num(wr.get('sma')),'sma50':_num(wr.get('sma50')),'as_of':TODAY,'source':'STOCKHISTORY','dq':None}
    if r['price'] is None: return None
    return _derive(r)
def build_tab(tab_key, qmap, wb_rows, prior_tab):
    built=[]; stats={'n':len(wb_rows),'fmp':0,'agree':0,'wins':0,'diverge':0,'suspect':[],'gaps':[],'carried':[],'shist':[]}
    for wr in wb_rows:
        tk=wr['ticker']; q=qmap.get(tk.upper()) or qmap.get(tk.upper().replace('.','-'))
        row=derive_row(tk,q) if q else None
        if row is None:
            shsev,_=ig._classify({'price':_num(wr.get('price')),'high':_num(wr.get('high')),'sma':_num(wr.get('sma')),'sma50':_num(wr.get('sma50'))})
            shrow=derive_from_wb(tk,wr)
            if shrow is not None and shsev!='CORRUPT' and not (shsev and shsev.startswith('MISSING (price/high)')):
                if shsev and shsev.startswith('MISSING'): shrow['dq']='missing'
                built.append(shrow); stats['shist'].append(tk); continue
            p=(prior_tab or {}).get(tk)
            if p and _num(p.get('price')) is not None:
                cf={k:p.get(k) for k in OUT_COLS if k in p}; cf['ticker']=tk; cf['source']='carry-forward'; cf['dq']='stale'
                built.append(cf); stats['carried'].append(tk)
            else:
                built.append({'ticker':tk,'source':'none','dq':'missing'}); stats['gaps'].append(tk)
            continue
        stats['fmp']+=1; sev,_=ig._classify(row)
        if sev=='CORRUPT':
            row['dq']='error'; stats['suspect'].append(tk); built.append(row); continue
        built.append(row)
        wsev,_=ig._classify({'price':_num(wr.get('price')),'high':_num(wr.get('high')),'sma':_num(wr.get('sma')),'sma50':_num(wr.get('sma50'))})
        wp=_num(wr.get('price')); ws_=_num(wr.get('sma'))
        if wsev=='CORRUPT' or (wsev and wsev.startswith('MISSING')): stats['wins']+=1
        elif wp and abs(row['price']/wp-1)<=0.02 and (not ws_ or not row.get('sma') or abs(row['sma']/ws_-1)<=0.03): stats['agree']+=1
        else: stats['diverge']+=1
    return built, stats
def shape_data_row(m, wr, is_fund, snap):
    wr = wr or {}; sn = (snap or {}).get(m['ticker'], {})
    d = {'ticker': m['ticker'], 'company': wr.get('company')}
    d['category' if is_fund else 'sector'] = wr.get('category' if is_fund else 'sector')
    for k in ['price','high','low','off','status','signal','recover','sma','smapct','sma50','sma50pct','cross','dq']:
        d[k] = m.get(k)
    if is_fund:
        d['divyield'] = sn.get('divyield', wr.get('divyield'))
    else:
        d['consensus'] = sn.get('consensus', wr.get('consensus'))
        d['fwdpe'] = sn.get('fwdpe', wr.get('fwdpe'))
    return d
def write_xlsx(all_rows, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    for tab in TABS:
        ws=wb.create_sheet(tab); ws.append(OUT_COLS)
        for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F3864')
        for row in all_rows[tab]: ws.append([row.get(k) for k in OUT_COLS])
        ws.freeze_panes='A2'
    wb.save(path)
def main():
    fmp_path, tab_arg = sys.argv[1], sys.argv[2]
    wb_path=WB_DEFAULT; md_path=MD_DEFAULT; do_write='--write' in sys.argv
    emit=sys.argv[sys.argv.index('--emit-data')+1] if '--emit-data' in sys.argv else None
    snap=json.load(open(sys.argv[sys.argv.index('--snap')+1])) if '--snap' in sys.argv else {}
    if '--wb' in sys.argv: wb_path=sys.argv[sys.argv.index('--wb')+1]
    if '--prior' in sys.argv: md_path=sys.argv[sys.argv.index('--prior')+1]
    quotes=json.load(open(fmp_path)); qmap={str(q.get('symbol')).strip().upper():q for q in quotes if q.get('symbol')}
    # merge --extra (e.g. Bigdata-sourced) quotes in the same FMP shape
    if '--extra' in sys.argv:
        for q in json.load(open(sys.argv[sys.argv.index('--extra')+1])):
            if q.get('symbol'): qmap[str(q['symbol']).strip().upper()]=q
    import openpyxl
    wb=openpyxl.load_workbook(wb_path,data_only=True,read_only=True); prior=load_prior(md_path)
    tabs=TABS if tab_arg.upper()=='ALL' else [tab_arg]
    all_rows={}; wbmaps={}; agg={'n':0,'fmp':0,'agree':0,'wins':0,'diverge':0,'suspect':[],'gaps':[],'carried':[],'shist':[]}
    print(f"=== build_market_data — {'ALL tabs' if tab_arg.upper()=='ALL' else tab_arg} ===")
    print(f"{'tab':<12}{'n':>5}{'FMP':>6}{'agree':>7}{'FMPfix':>8}{'diverg':>8}{'suspct':>8}{'gap':>5}{'carry':>6}")
    for tab in tabs:
        wb_rows=read_wb_tab(wb,tab); rows,s=build_tab(tab,qmap,wb_rows,prior.get(tab))
        all_rows[tab]=rows; wbmaps[tab]={wr['ticker']:wr for wr in wb_rows}
        print(f"{tab:<12}{s['n']:>5}{s['fmp']:>6}{s['agree']:>7}{s['wins']:>8}{s['diverge']:>8}{len(s['suspect']):>8}{len(s['gaps']):>5}{len(s['carried']):>6}")
        for k in ('n','fmp','agree','wins','diverge'): agg[k]+=s[k]
        for k in ('suspect','gaps','carried','shist'): agg[k]+=s[k]
    if len(tabs)>1:
        print(f"{'TOTAL':<12}{agg['n']:>5}{agg['fmp']:>6}{agg['agree']:>7}{agg['wins']:>8}{agg['diverge']:>8}{len(agg['suspect']):>8}{len(agg['gaps']):>5}{len(agg['carried']):>6}")
    print(f"\nFMP-suspect (band-aware guard flagged): {agg['suspect'] or 'none'}")
    print(f"STOCKHISTORY tail-rung filled: {len(agg['shist'])} -> {sorted(agg['shist'])}")
    print(f"TRUE gaps (missing everywhere): {len(agg['gaps'])} -> {sorted(agg['gaps'])}")
    ev("INFO","build","universe built", n=agg['n'], fmp=agg['fmp'], shist=len(agg['shist']), carry=len(agg['carried']), diverge=agg['diverge'], suspect=len(agg['suspect']), gaps=len(agg['gaps']))
    if agg['gaps']: ev("WARN","build","true gaps (missing everywhere)", tickers=sorted(agg['gaps']))
    if agg['suspect']: ev("WARN","guard","FMP band-guard flagged suspects", tickers=agg['suspect'])
    if do_write:
        write_xlsx(all_rows,md_path); print(f"\nwrote {sum(len(v) for v in all_rows.values())} rows / {len(all_rows)} tabs -> {os.path.relpath(md_path,HERE)}")
    if emit:
        data={tab:[shape_data_row(m,wbmaps[tab].get(m['ticker']),tab in FUND_KEYS,snap) for m in all_rows[tab]] for tab in tabs}
        json.dump(data,open(emit,'w'),separators=(',',':')); print(f"emitted DATA ({sum(len(v) for v in data.values())} rows) -> {os.path.relpath(emit,HERE)}")
if __name__=='__main__': main()
